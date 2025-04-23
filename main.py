import logging
import json
import io
from datetime import datetime, timedelta
from collections import defaultdict

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
import aiohttp
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleRequest

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

# === Settings ===
API_VERSION     = "v17"
DEVELOPER_TOKEN = "D4yv61IQ8R0JaE5dxrd1Uw"
CLIENT_ID       = "167266694231-g7hvta57r99etbp3sos3jfi7q7h4ef44.apps.googleusercontent.com"
CLIENT_SECRET   = "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"

logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)

# ───────── Helpers for API calls ─────────

async def get_access_token(refresh_token: str) -> str:
    try:
        creds = Credentials(
            token=None, refresh_token=refresh_token,
            token_uri="https://oauth2.googleapis.com/token",
            client_id=CLIENT_ID, client_secret=CLIENT_SECRET
        )
        creds.refresh(GoogleRequest())
        return creds.token
    except Exception as e:
        logging.error("Google token refresh failed: %s", e)
        raise HTTPException(401, f"Falha ao renovar token Google: {e}")

async def discover_customer_id(token: str) -> str:
    url = f"https://googleads.googleapis.com/{API_VERSION}/customers:listAccessibleCustomers"
    headers = {"Authorization": f"Bearer {token}", "developer-token": DEVELOPER_TOKEN}
    async with aiohttp.ClientSession() as sess:
        async with sess.get(url, headers=headers) as resp:
            text = await resp.text()
            if resp.status != 200:
                raise HTTPException(502, f"Google listAccessibleCustomers: {text}")
            names = json.loads(text).get("resourceNames", [])
            if not names:
                raise HTTPException(502, "Google: sem customers acessíveis")
            return names[0].split("/")[-1]

async def google_ads_list(refresh_token: str, with_trends: bool = False):
    token = await get_access_token(refresh_token)
    cid   = await discover_customer_id(token)
    url   = f"https://googleads.googleapis.com/{API_VERSION}/customers/{cid}/googleAds:search"
    headers = {"Authorization": f"Bearer {token}", "developer-token": DEVELOPER_TOKEN}

    # Buscar budget e métricas
    q_active = """
        SELECT campaign.id, campaign.name, campaign.status,
               campaign_budget.amount_micros,
               metrics.impressions, metrics.clicks, metrics.cost_micros,
               metrics.average_cpc, metrics.ctr, metrics.conversions
        FROM campaign
        WHERE campaign.status = 'ENABLED'
    """
    async with aiohttp.ClientSession() as sess:
        async with sess.post(url, headers=headers, json={"query": q_active}) as resp:
            text = await resp.text()
            if resp.status != 200:
                raise HTTPException(resp.status, f"Google search: {text}")
            results = json.loads(text).get("results", [])

    rows = []
    for r in results:
        budget = r.get("campaignBudget", {}).get("amountMicros", 0) / 1e6
        spend  = r["metrics"].get("costMicros", 0) / 1e6
        clicks  = int(r["metrics"].get("clicks", 0))
        impr    = int(r["metrics"].get("impressions", 0))
        conv    = int(r["metrics"].get("conversions", 0))
        rows.append({
            "Campaign ID": r["campaign"]["id"],
            "Name":        r["campaign"]["name"],
            "Budget":      round(budget, 2),
            "Spend":       round(spend, 2),
            "Impressions": impr,
            "Clicks":      clicks,
            "CTR (%)":     round(r["metrics"].get("ctr", 0)*100, 2),
            "CPC":         round(r["metrics"].get("averageCpc", 0)/1e6, 2),
            "Conversions": conv,
            "CPA":         round(spend / max(conv,1), 2)
        })
    df = pd.DataFrame(rows)

    trends, changes = None, {}
    if with_trends:
        # Insights últimos 14 dias
        q_trend = """
            SELECT segments.date, metrics.impressions, metrics.clicks
            FROM campaign
            WHERE campaign.status = 'ENABLED'
              AND segments.date DURING LAST_14_DAYS
        """
        async with aiohttp.ClientSession() as sess:
            async with sess.post(url, headers=headers, json={"query": q_trend}) as resp:
                text = await resp.text()
                if resp.status != 200:
                    raise HTTPException(resp.status, f"Google trend: {text}")
                results = json.loads(text).get("results", [])

        by_date = defaultdict(lambda: {"Impressions":0,"Clicks":0})
        for r in results:
            d = r["segments"]["date"]
            by_date[d]["Impressions"] += int(r["metrics"]["impressions"])
            by_date[d]["Clicks"]      += int(r["metrics"]["clicks"])
        dates = sorted(by_date)
        last7 = dates[-7:]
        prev7 = dates[-14:-7] if len(dates)>=14 else dates[:max(len(dates)-7,0)]
        curr_imp = sum(by_date[d]["Impressions"] for d in last7)
        prev_imp = sum(by_date[d]["Impressions"] for d in prev7) if prev7 else 0
        curr_clk = sum(by_date[d]["Clicks"] for d in last7)
        prev_clk = sum(by_date[d]["Clicks"] for d in prev7) if prev7 else 0
        changes = {
            "Impr Δ (%)": round((curr_imp - prev_imp)/max(prev_imp,1)*100,2),
            "Clk Δ (%)":  round((curr_clk - prev_clk)/max(prev_clk,1)*100,2)
        }
        trends = pd.DataFrame({
            "Date":        last7,
            "Impressions": [by_date[d]["Impressions"] for d in last7],
            "Clicks":      [by_date[d]["Clicks"] for d in last7]
        })

    return df, trends, changes

async def meta_ads_list(refresh_token: str, account_id: str, with_trends: bool = False):
    # Campaigns com budget
    url_c = f"https://graph.facebook.com/v16.0/act_{account_id}/campaigns"
    params = {
        "fields":"id,name,status,amount_spent,daily_budget,lifetime_budget",
        "access_token": refresh_token
    }
    async with aiohttp.ClientSession() as sess:
        async with sess.get(url_c, params=params) as resp:
            text = await resp.text()
            if resp.status != 200:
                raise HTTPException(resp.status, f"Meta campaigns: {text}")
            data = json.loads(text).get("data", [])

    # Insights últimos 14 dias (para Δ e tendências)
    since_14 = (datetime.now().date() - timedelta(days=14)).isoformat()
    until    = datetime.now().date().isoformat()
    ins_url = f"https://graph.facebook.com/v16.0/act_{account_id}/insights"
    ins_params = {
        "level":"campaign",
        "fields":"campaign_id,impressions,clicks,spend,actions",
        "time_range": json.dumps({"since": since_14, "until": until}),
        "access_token": refresh_token
    }
    async with aiohttp.ClientSession() as sess:
        async with sess.get(ins_url, params=ins_params) as resp:
            text = await resp.text()
            if resp.status != 200:
                raise HTTPException(resp.status, f"Meta insights: {text}")
            insights = json.loads(text).get("data", [])

    # Agrupar insights por campanha
    map_ins = {i["campaign_id"]: i for i in insights}
    rows = []
    for c in data:
        spent = float(c.get("amount_spent",0))
        # orçamento: lifetime se existir, senão daily*30
        raw_budget = c.get("lifetime_budget") or str(int(c.get("daily_budget",0))*30)
        budget = float(raw_budget)/100
        ins = map_ins.get(c["id"], {})
        clicks = int(ins.get("clicks",0))
        impr   = int(ins.get("impressions",0))
        # conversões = soma de todas as actions
        conv = sum(int(a.get("value",0)) for a in ins.get("actions",[]))
        rows.append({
            "Campaign ID": c["id"],
            "Name":        c["name"],
            "Budget":      round(budget,2),
            "Spend":       round(spent,2),
            "Impressions": impr,
            "Clicks":      clicks,
            "CTR (%)":     round(clicks/max(impr,1)*100,2),
            "CPC":         round(spent/max(clicks,1),2),
            "Conversions": conv,
            "CPA":         round(spent/max(conv,1),2)
        })
    df = pd.DataFrame(rows)

    trends, changes = None, {}
    if with_trends:
        by_date = defaultdict(lambda: {"Impressions":0,"Clicks":0})
        for i in insights:
            dt = i["date_start"]
            by_date[dt]["Impressions"] += int(i.get("impressions",0))
            by_date[dt]["Clicks"]      += int(i.get("clicks",0))
        dates = sorted(by_date)
        last7 = dates[-7:]
        prev7 = dates[-14:-7] if len(dates)>=14 else dates[:max(len(dates)-7,0)]
        curr_imp = sum(by_date[d]["Impressions"] for d in last7)
        prev_imp = sum(by_date[d]["Impressions"] for d in prev7) if prev7 else 0
        curr_clk = sum(by_date[d]["Clicks"] for d in last7)
        prev_clk = sum(by_date[d]["Clicks"] for d in prev7) if prev7 else 0
        changes = {
            "Impr Δ (%)": round((curr_imp - prev_imp)/max(prev_imp,1)*100,2),
            "Clk Δ (%)":  round((curr_clk - prev_clk)/max(prev_clk,1)*100,2)
        }
        trends = pd.DataFrame({
            "Date":        last7,
            "Impressions": [by_date[d]["Impressions"] for d in last7],
            "Clicks":      [by_date[d]["Clicks"] for d in last7]
        })

    return df, trends, changes

def make_xlsx(df: pd.DataFrame, trends: pd.DataFrame, changes: dict) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        wb = writer.book
        ws = wb.create_sheet("Dashboard")

        # configurações visuais
        primary, secondary = "7F56D9", "E839BC"
        header_font = Font(bold=True, color="FFFFFF")
        val_font    = Font(bold=True, size=14, color="FFFFFF")
        align = Alignment(horizontal="center", vertical="center")

        # ─── KPI boxes ───
        kpis = {
            "Total Budget":       round(df["Budget"].sum(),2),
            "Total Spend":        round(df["Spend"].sum(),2),
            "+ / -":              round(df["Budget"].sum() - df["Spend"].sum(),2),
            "Impressions":        int(df["Impressions"].sum()),
            "Clicks":             int(df["Clicks"].sum()),
            "Conversions":        int(df["Conversions"].sum()),
            "Cost/Click":         round(df["Spend"].sum()/max(df["Clicks"].sum(),1),2),
            "Cost/Conv":          round(df["Spend"].sum()/max(df["Conversions"].sum(),1),2),
            "CTR (%)":            round(df["Clicks"].sum()/max(df["Impressions"].sum(),1)*100,2),
            "CPA (Avg)":          round(df["Spend"].sum()/max(df["Conversions"].sum(),1),2),
            **changes
        }
        col = 1
        for title, value in kpis.items():
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+2)
            c1 = ws.cell(row=1, column=col, value=title)
            c1.font = header_font; c1.fill = PatternFill("solid", fgColor=primary); c1.alignment = align
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+2)
            c2 = ws.cell(row=2, column=col, value=value)
            c2.font = val_font; c2.fill = PatternFill("solid", fgColor=secondary); c2.alignment = align
            col += 3

        # ─── Escrever dados de campanha (oculto sob charts) ───
        start_data_row = 5
        for ci, col_name in enumerate(df.columns, start=1):
            ws.cell(row=start_data_row, column=ci, value=col_name)
        for ri, row in enumerate(df.itertuples(index=False), start=start_data_row+1):
            for ci, val in enumerate(row, start=1):
                ws.cell(row=ri, column=ci, value=val)
        max_row = start_data_row + len(df)
        # ─── Budget vs Spend per Campaign ───
        name_idx   = df.columns.get_loc("Name") + 1
        budget_idx = df.columns.get_loc("Budget") + 1
        spend_idx  = df.columns.get_loc("Spend") + 1

        chart1 = BarChart()
        chart1.title = "Budget vs Spend per Campaign"
        chart1.y_axis.title = "Value"
        data = Reference(ws, min_col=budget_idx, min_row=start_data_row,
                         max_col=spend_idx, max_row=max_row)
        cats = Reference(ws, min_col=name_idx, min_row=start_data_row+1,
                         max_row=max_row)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.series[0].graphicalProperties.solidFill = primary
        chart1.series[1].graphicalProperties.solidFill = secondary
        ws.add_chart(chart1, "A8")

        # ─── CPA per Campaign ───
        cpa_idx = df.columns.get_loc("CPA") + 1
        chart2 = BarChart()
        chart2.title = "CPA per Campaign"
        chart2.x_axis.title = ""
        chart2.y_axis.title = "CPA"
        data2 = Reference(ws, min_col=cpa_idx, min_row=start_data_row,
                          max_row=max_row)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.series[0].graphicalProperties.solidFill = primary
        ws.add_chart(chart2, "H8")

        # ─── CPC per Campaign ───
        cpc_idx = df.columns.get_loc("CPC") + 1
        chart3 = BarChart()
        chart3.title = "CPC per Campaign"
        data3 = Reference(ws, min_col=cpc_idx, min_row=start_data_row,
                          max_row=max_row)
        chart3.add_data(data3, titles_from_data=True)
        chart3.set_categories(cats)
        chart3.series[0].graphicalProperties.solidFill = primary
        ws.add_chart(chart3, "N8")

        # ─── Acquisitions (Pie) ───
        conv_idx = df.columns.get_loc("Conversions") + 1
        pie = PieChart()
        pie.title = "Acquisitions"
        data4 = Reference(ws, min_col=conv_idx, min_row=start_data_row,
                          max_row=max_row)
        labels4 = Reference(ws, min_col=name_idx, min_row=start_data_row+1,
                             max_row=max_row)
        pie.add_data(data4, titles_from_data=True)
        pie.set_categories(labels4)
        ws.add_chart(pie, "H1")

        # ─── CTR per Campaign (Horizontal Bar) ───
        ctr_idx = df.columns.get_loc("CTR (%)") + 1
        chart5 = BarChart(orientation="bar")
        chart5.title = "CTR per Campaign"
        data5 = Reference(ws, min_col=ctr_idx, min_row=start_data_row,
                          max_row=max_row)
        chart5.add_data(data5, titles_from_data=True)
        chart5.set_categories(labels4)
        chart5.series[0].graphicalProperties.solidFill = primary
        ws.add_chart(chart5, "N1")

    return buf.getvalue()

# ───────── Endpoints XLSX ─────────

@app.get("/export_google_active_campaigns_csv")
async def export_google_xlsx(google_refresh_token: str = Query(..., alias="google_refresh_token")):
    df, trends, changes = await google_ads_list(google_refresh_token, with_trends=True)
    xlsx = make_xlsx(df, trends, changes)
    return JSONResponse({
        "fileName": "google_active_campaigns.xlsx",
        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "bytes": list(xlsx)
    })

@app.get("/export_meta_active_campaigns_csv")
async def export_meta_xlsx(
    meta_account_id:   str = Query(..., alias="meta_account_id"),
    meta_access_token: str = Query(..., alias="meta_access_token")
):
    df, trends, changes = await meta_ads_list(meta_access_token, meta_account_id, with_trends=True)
    xlsx = make_xlsx(df, trends, changes)
    return JSONResponse({
        "fileName": "meta_active_campaigns.xlsx",
        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "bytes": list(xlsx)
    })

@app.get("/export_combined_active_campaigns_csv")
async def export_combined_xlsx(
    google_refresh_token: str = Query(..., alias="google_refresh_token"),
    meta_account_id:     str = Query(..., alias="meta_account_id"),
    meta_access_token:   str = Query(..., alias="meta_access_token")
):
    g_df, g_tr, g_ch = await google_ads_list(google_refresh_token, with_trends=True)
    m_df, m_tr, m_ch = await meta_ads_list(meta_access_token, meta_account_id, with_trends=True)
    df = pd.concat([g_df, m_df], ignore_index=True)
    # combinar Δ como média
    combined_changes = {
        "Impr Δ (%)": round((g_ch.get("Impr Δ (%)",0) + m_ch.get("Impr Δ (%)",0))/2, 2),
        "Clk Δ (%)":  round((g_ch.get("Clk Δ (%)",0) + m_ch.get("Clk Δ (%)",0))/2, 2)
    }
    # tendências não usadas no combined chart
    xlsx = make_xlsx(df, None, combined_changes)
    return JSONResponse({
        "fileName": "combined_active_campaigns.xlsx",
        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "bytes": list(xlsx)
    })

if __name__ == "__main__":
    logging.info("Starting export service on port 8080")
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8080, reload=True)
